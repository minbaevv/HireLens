"""Сервис экспорта: CSV/Excel список кандидатов + PDF (отчёт кандидата и список)."""
import csv
import io
import json
import logging
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from typing import List, Optional

logger = logging.getLogger(__name__)

from app.core.config import settings as _settings


def _local_dt(dt):
    """UTC-наивный/aware datetime → локальное время SCHEDULING_TIMEZONE (Бишкек)."""
    if not dt:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    try:
        return dt.astimezone(ZoneInfo(_settings.SCHEDULING_TIMEZONE))
    except Exception:
        return dt


def _tags_str(raw) -> str:
    """JSON-строку тегов кандидата превращает в 'tag1, tag2'."""
    if not raw:
        return ""
    try:
        val = json.loads(raw) if isinstance(raw, str) else raw
        if isinstance(val, list):
            return ", ".join(str(x) for x in val if x)
    except Exception:
        pass
    return str(raw)


# Общий набор колонок для табличного экспорта (CSV/Excel)
_EXPORT_HEADERS = [
    "ID", "Имя", "Email", "Телефон", "Статус", "Оценка",
    "Пре-скрин", "Рекомендация", "Теги", "Вакансия ID", "Дата подачи",
]


def _row_values(c) -> list:
    """Значения одной строки кандидата (строковые, для CSV)."""
    return [
        c.id,
        c.name,
        c.email,
        getattr(c, "phone", None) or "",
        c.status.value if hasattr(c.status, "value") else str(c.status),
        f"{c.score:.1f}" if c.score is not None else "",
        f"{c.pre_score:.0f}" if getattr(c, "pre_score", None) is not None else "",
        c.recommendation or "",
        _tags_str(getattr(c, "tags", None)),
        c.job_id,
        _local_dt(c.created_at).strftime("%Y-%m-%d %H:%M") if c.created_at else "",
    ]


def generate_candidates_csv(candidates: list) -> bytes:
    """Генерирует CSV со списком кандидатов."""
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow(_EXPORT_HEADERS)
    for c in candidates:
        writer.writerow(_row_values(c))
    return output.getvalue().encode("utf-8-sig")  # utf-8-sig для Excel


def generate_candidates_xlsx(candidates: list) -> bytes:
    """Генерирует Excel (.xlsx) со списком кандидатов."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("Библиотека openpyxl не установлена. Запусти: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Кандидаты"

    header_fill = PatternFill(fill_type="solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(_EXPORT_HEADERS)
    for col in range(1, len(_EXPORT_HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for c in candidates:
        row = _row_values(c)
        # Оценка и пре-скрин — числами, чтобы Excel корректно сортировал
        # (индексы сдвинуты на +1 после добавления колонки «Телефон»)
        if row[5] != "":
            try:
                row[5] = float(row[5])
            except Exception:
                pass
        if row[6] != "":
            try:
                row[6] = int(float(row[6]))
            except Exception:
                pass
        ws.append(row)

    widths = [6, 24, 30, 18, 14, 8, 10, 16, 26, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(_EXPORT_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{max(1, ws.max_row)}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _register_cyrillic_font():
    """Регистрирует шрифт с поддержкой кириллицы."""
    import os
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Пути к DejaVuSans на разных ОС
    font_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",       # Linux (Debian/Ubuntu)
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",                 # Linux (Alpine)
        "/usr/share/fonts/TTF/DejaVuSans.ttf",                    # Arch Linux
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",   # macOS
        "C:/Windows/Fonts/arial.ttf",                             # Windows
    ]

    for path in font_paths:
        if os.path.exists(path):
            pdfmetrics.registerFont(TTFont("CyrillicFont", path))
            bold_path = path.replace("DejaVuSans.ttf", "DejaVuSans-Bold.ttf")
            if os.path.exists(bold_path):
                pdfmetrics.registerFont(TTFont("CyrillicFont-Bold", bold_path))
            else:
                pdfmetrics.registerFont(TTFont("CyrillicFont-Bold", path))
            return "CyrillicFont"

    # Fallback: скачиваем DejaVuSans если нет локально
    try:
        import urllib.request
        import tempfile
        url = "https://github.com/dejavu-fonts/dejavu-fonts/raw/master/ttf/DejaVuSans.ttf"
        tmp = tempfile.mktemp(suffix=".ttf")
        urllib.request.urlretrieve(url, tmp)
        pdfmetrics.registerFont(TTFont("CyrillicFont", tmp))
        pdfmetrics.registerFont(TTFont("CyrillicFont-Bold", tmp))
        return "CyrillicFont"
    except Exception:
        pass

    return None  # шрифт не найден, используем Helvetica (без кириллицы)


def generate_candidate_pdf(candidate, job_title: str, messages: list = None) -> bytes:
    """Генерирует PDF отчёт по кандидату."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
    except ImportError:
        raise RuntimeError("Библиотека reportlab не установлена. Запусти: pip install reportlab")

    # Регистрируем кириллический шрифт
    font_name = _register_cyrillic_font() or "Helvetica"
    font_bold = (font_name + "-Bold") if font_name != "Helvetica" else "Helvetica-Bold"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    # Стили с поддержкой кириллицы
    title_style = ParagraphStyle(
        "CyrTitle", parent=styles["Title"],
        fontSize=18, spaceAfter=12, textColor=colors.HexColor("#4F46E5"),
        fontName=font_name,
    )
    heading_style = ParagraphStyle(
        "CyrHeading", parent=styles["Heading2"],
        fontSize=13, spaceBefore=16, spaceAfter=6, textColor=colors.HexColor("#1F2937"),
        fontName=font_bold,
    )
    normal_style = ParagraphStyle(
        "CyrNormal", parent=styles["Normal"],
        fontSize=10, spaceAfter=4, leading=14,
        fontName=font_name,
    )
    label_style = ParagraphStyle(
        "CyrLabel", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#6B7280"),
        fontName=font_name,
    )

    # Цвет рекомендации
    rec_colors = {
        "hire": colors.HexColor("#10B981"),
        "maybe": colors.HexColor("#F59E0B"),
        "reject": colors.HexColor("#EF4444"),
    }
    rec_labels = {"hire": "HIRE", "maybe": "MAYBE", "reject": "REJECT"}
    rec = candidate.recommendation or "maybe"
    rec_color = rec_colors.get(rec, colors.gray)
    rec_label = rec_labels.get(rec, rec.upper())

    story = []

    # Заголовок
    story.append(Paragraph("HireLens — Отчёт кандидата", title_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E5E7EB")))
    story.append(Spacer(1, 0.4 * cm))

    # Основная инфо
    info_data = [
        ["Кандидат:", candidate.name],
        ["Email:", candidate.email],
        ["Телефон:", getattr(candidate, "phone", None) or "—"],
        ["Вакансия:", job_title],
        ["Статус:", candidate.status.value.upper()],
        ["Дата:", _local_dt(candidate.created_at).strftime("%d.%m.%Y") if candidate.created_at else "—"],
    ]
    info_table = Table(info_data, colWidths=[4 * cm, 13 * cm])
    info_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#6B7280")),
        ("TEXTCOLOR", (1, 0), (1, -1), colors.HexColor("#111827")),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTNAME", (1, 0), (1, -1), font_bold),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.4 * cm))

    # Скоринг + рекомендация
    if candidate.score is not None:
        story.append(Paragraph("📊 Результат AI-интервью", heading_style))
        score_data = [
            ["Оценка:", f"{candidate.score:.0f} / 100"],
            ["Рекомендация:", rec_label],
        ]
        score_table = Table(score_data, colWidths=[4 * cm, 13 * cm])
        score_table.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#6B7280")),
            ("TEXTCOLOR", (1, 0), (1, 0), colors.HexColor("#4F46E5")),
            ("TEXTCOLOR", (1, 1), (1, 1), rec_color),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTNAME", (1, 0), (1, -1), font_bold),
            ("FONTSIZE", (1, 0), (1, 0), 20),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(score_table)

    # Резюме
    if candidate.summary:
        story.append(Paragraph("📝 Резюме AI", heading_style))
        story.append(Paragraph(candidate.summary, normal_style))

    # Резюме кандидата
    if candidate.resume_text:
        story.append(Paragraph("📄 Резюме кандидата", heading_style))
        # Ограничиваем длину резюме в PDF
        resume_text = candidate.resume_text[:2000]
        if len(candidate.resume_text) > 2000:
            resume_text += "..."
        story.append(Paragraph(resume_text.replace("\n", "<br/>"), normal_style))

    # Транскрипт интервью
    if messages:
        story.append(Paragraph("🎤 Транскрипт интервью", heading_style))
        for msg in messages:
            role = "AI" if msg.role.value == "ai" else "Кандидат"
            color = "#4F46E5" if msg.role.value == "ai" else "#111827"
            story.append(Paragraph(
                f'<font color="{color}"><b>{role}:</b></font> {msg.content[:500]}',
                normal_style,
            ))
            story.append(Spacer(1, 0.2 * cm))

    # Футер
    story.append(Spacer(1, 1 * cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E5E7EB")))
    story.append(Paragraph(
        f"Сгенерировано: {_local_dt(datetime.now(timezone.utc)).strftime('%d.%m.%Y %H:%M')} | HireLens",
        ParagraphStyle("CyrFooter", parent=styles["Normal"], fontSize=8,
                       textColor=colors.HexColor("#9CA3AF"), spaceBefore=6,
                       fontName=font_name),
    ))

    doc.build(story)
    return buffer.getvalue()


def generate_candidates_list_pdf(candidates: list, company_name: str = None) -> bytes:
    """Генерирует PDF-отчёт со списком кандидатов (таблица, альбомная ориентация)."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
    except ImportError:
        raise RuntimeError("Библиотека reportlab не установлена. Запусти: pip install reportlab")

    font_name = _register_cyrillic_font() or "Helvetica"
    font_bold = (font_name + "-Bold") if font_name != "Helvetica" else "Helvetica-Bold"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CyrListTitle", parent=styles["Title"],
        fontSize=16, spaceAfter=8, textColor=colors.HexColor("#4F46E5"),
        fontName=font_name,
    )
    sub_style = ParagraphStyle(
        "CyrListSub", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#6B7280"),
        fontName=font_name, spaceAfter=8,
    )
    cell_style = ParagraphStyle(
        "CyrCell", parent=styles["Normal"],
        fontSize=8, leading=10, fontName=font_name,
    )
    head_cell_style = ParagraphStyle(
        "CyrHeadCell", parent=styles["Normal"],
        fontSize=8, leading=10, fontName=font_bold, textColor=colors.white,
    )

    def P(txt, style=cell_style):
        return Paragraph(str(txt) if txt is not None else "", style)

    story = []
    header_txt = "HireLens — Список кандидатов"
    if company_name:
        header_txt += f" · {company_name}"
    story.append(Paragraph(header_txt, title_style))
    story.append(Paragraph(
        f"Всего: {len(candidates)} · Сгенерировано: "
        f"{_local_dt(datetime.now(timezone.utc)).strftime('%d.%m.%Y %H:%M')}",
        sub_style,
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E5E7EB")))
    story.append(Spacer(1, 0.3 * cm))

    headers = ["ID", "Имя", "Email", "Статус", "Оценка", "Пре-скрин", "Рекоменд.", "Теги", "Дата"]
    table_data = [[P(h, head_cell_style) for h in headers]]
    for c in candidates:
        table_data.append([
            P(c.id),
            P(c.name),
            P(c.email),
            P(c.status.value if hasattr(c.status, "value") else c.status),
            P(f"{c.score:.0f}" if c.score is not None else "—"),
            P(f"{c.pre_score:.0f}" if getattr(c, "pre_score", None) is not None else "—"),
            P((c.recommendation or "—").upper()),
            P(_tags_str(getattr(c, "tags", None)) or "—"),
            P(_local_dt(c.created_at).strftime("%d.%m.%Y") if c.created_at else "—"),
        ])

    col_widths = [1.2 * cm, 4.5 * cm, 6 * cm, 2.5 * cm, 1.6 * cm, 1.8 * cm, 2.4 * cm, 4 * cm, 2.2 * cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
    ]))
    story.append(table)

    if not candidates:
        story.append(Spacer(1, 0.5 * cm))
        story.append(Paragraph("Нет кандидатов для экспорта.", sub_style))

    doc.build(story)
    return buffer.getvalue()
